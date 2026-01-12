#!/usr/bin/env python3
"""
Generate an SR report for the CT-axi-postop study based on labels defined
in a plain-text file. The generated report mirrors the structure of the
template SR stored under `scripts/SR Saved report/`.

Also supports generating SR reports from the `refdata-template` sheet in an
XLSX workbook. In that mode, the XLSX describes which annotations should
exist per case+image; coordinates are generated as deterministic placeholders
until a doctor annotates the real positions.
"""

from __future__ import annotations

import argparse
import copy
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pydicom
from pydicom.dataset import Dataset
from pydicom.uid import generate_uid


SR_STORAGE_UID = "1.2.840.10008.5.1.4.1.1.88.22"  # Enhanced SR
POINT_CODE = ("111010", "DCM", "Center")
MEASUREMENT_GROUP_CODE = ("125007", "DCM", "Measurement Group")
TRACKING_IDENTIFIER_CODE = ("112039", "DCM", "Tracking Identifier")
TRACKING_UID_CODE = ("112040", "DCM", "Tracking Unique Identifier")
TRACKING_IDENTIFIER_TEXT = "Cornerstone3DTools@^0.1.0:Probe:CustomProbe"
IMAGING_MEASUREMENTS_CODE = ("126010", "DCM", "Imaging Measurements")
IMAGE_LIBRARY_CODE = ("111028", "DCM", "Image Library")
IMAGE_LIBRARY_GROUP_CODE = ("126200", "DCM", "Image Library Group")
LANGUAGE_CODE = ("121049", "DCM", "Language of Content Item and Descendants")
COUNTRY_OF_LANGUAGE_CODE = ("121046", "DCM", "Country of Language")
LANGUAGE_ENGLISH = ("eng", "RFC5646", "English")
COUNTRY_US = ("US", "ISO3166_1", "United States")
OBSERVER_CODE = ("121008", "DCM", "Person Observer Name")
PROCEDURE_REPORTED_CODE = ("121058", "DCM", "Procedure reported")
UNKNOWN_PROCEDURE_CODE = ("1", "99dcmjs", "Unknown procedure")
FINDING_CODE = ("121071", "DCM", "Finding")
FINDING_SITE_CODE = ("363698007", "SCT", "Finding Site")
HIDDEN_UNANNOTATED_CODE = ("HIDDEN", "99MEDICALVIEWER", "")
CORNERSTONE_FREETEXT = ("CORNERSTONEFREETEXT", "CORNERSTONEJS")

DEFAULT_XLSX_SHEET = "refdata-template"


def _make_code(code_value: str, scheme: str, meaning: str) -> Dataset:
    ds = Dataset()
    ds.CodeValue = code_value
    ds.CodingSchemeDesignator = scheme
    ds.CodeMeaning = meaning
    return ds


def _make_language_item() -> Dataset:
    lang_item = Dataset()
    lang_item.RelationshipType = "HAS CONCEPT MOD"
    lang_item.ValueType = "CODE"
    lang_item.ConceptNameCodeSequence = [Dataset()]
    lang_item.ConceptNameCodeSequence[0].CodeValue = LANGUAGE_CODE[0]
    lang_item.ConceptNameCodeSequence[0].CodingSchemeDesignator = LANGUAGE_CODE[1]
    lang_item.ConceptNameCodeSequence[0].CodeMeaning = LANGUAGE_CODE[2]
    lang_item.ConceptCodeSequence = [_make_code(*LANGUAGE_ENGLISH)]

    country = Dataset()
    country.RelationshipType = "HAS CONCEPT MOD"
    country.ValueType = "CODE"
    country.ConceptNameCodeSequence = [Dataset()]
    cn = country.ConceptNameCodeSequence[0]
    cn.CodeValue, cn.CodingSchemeDesignator, cn.CodeMeaning = COUNTRY_OF_LANGUAGE_CODE
    country.ConceptCodeSequence = [_make_code(*COUNTRY_US)]
    lang_item.ContentSequence = [country]
    return lang_item


def _make_observer_item(observer_name: str) -> Dataset:
    obs = Dataset()
    obs.RelationshipType = "HAS OBS CONTEXT"
    obs.ValueType = "PNAME"
    obs.ConceptNameCodeSequence = [Dataset()]
    cn = obs.ConceptNameCodeSequence[0]
    cn.CodeValue, cn.CodingSchemeDesignator, cn.CodeMeaning = OBSERVER_CODE
    obs.PersonName = observer_name
    return obs


def _make_procedure_item() -> Dataset:
    proc = Dataset()
    proc.RelationshipType = "HAS CONCEPT MOD"
    proc.ValueType = "CODE"
    proc.ConceptNameCodeSequence = [Dataset()]
    cn = proc.ConceptNameCodeSequence[0]
    cn.CodeValue, cn.CodingSchemeDesignator, cn.CodeMeaning = PROCEDURE_REPORTED_CODE
    proc.ConceptCodeSequence = [_make_code(*UNKNOWN_PROCEDURE_CODE)]
    return proc


def _make_image_library(image_reference: Dataset) -> Dataset:
    lib = Dataset()
    lib.RelationshipType = "CONTAINS"
    lib.ValueType = "CONTAINER"
    lib.ConceptNameCodeSequence = [Dataset()]
    cn = lib.ConceptNameCodeSequence[0]
    cn.CodeValue, cn.CodingSchemeDesignator, cn.CodeMeaning = IMAGE_LIBRARY_CODE
    lib.ContinuityOfContent = "SEPARATE"

    group = Dataset()
    group.RelationshipType = "CONTAINS"
    group.ValueType = "CONTAINER"
    group.ConceptNameCodeSequence = [Dataset()]
    gcn = group.ConceptNameCodeSequence[0]
    gcn.CodeValue, gcn.CodingSchemeDesignator, gcn.CodeMeaning = IMAGE_LIBRARY_GROUP_CODE
    group.ContinuityOfContent = "SEPARATE"

    image_item = Dataset()
    image_item.RelationshipType = "CONTAINS"
    image_item.ValueType = "IMAGE"
    image_item.ReferencedSOPSequence = [copy.deepcopy(image_reference)]
    group.ContentSequence = [image_item]
    lib.ContentSequence = [group]
    return lib


def _make_measurement_group(
    label: str,
    coords: Tuple[float, float],
    image_reference: Dataset,
) -> Dataset:
    group = Dataset()
    group.RelationshipType = "CONTAINS"
    group.ValueType = "CONTAINER"
    group.ConceptNameCodeSequence = [Dataset()]
    cn = group.ConceptNameCodeSequence[0]
    cn.CodeValue, cn.CodingSchemeDesignator, cn.CodeMeaning = MEASUREMENT_GROUP_CODE
    group.ContinuityOfContent = "SEPARATE"
    group.ContentSequence = []

    tracking = Dataset()
    tracking.RelationshipType = "HAS OBS CONTEXT"
    tracking.ValueType = "TEXT"
    tracking.ConceptNameCodeSequence = [Dataset()]
    tcn = tracking.ConceptNameCodeSequence[0]
    tcn.CodeValue, tcn.CodingSchemeDesignator, tcn.CodeMeaning = TRACKING_IDENTIFIER_CODE
    tracking.TextValue = TRACKING_IDENTIFIER_TEXT
    group.ContentSequence.append(tracking)

    tracking_uid = Dataset()
    tracking_uid.RelationshipType = "HAS OBS CONTEXT"
    tracking_uid.ValueType = "UIDREF"
    tracking_uid.ConceptNameCodeSequence = [Dataset()]
    tucn = tracking_uid.ConceptNameCodeSequence[0]
    tucn.CodeValue, tucn.CodingSchemeDesignator, tucn.CodeMeaning = TRACKING_UID_CODE
    tracking_uid.UID = generate_uid()
    group.ContentSequence.append(tracking_uid)

    # Finding site item used by the viewer to detect unannotated / hidden annotations.
    finding_site = Dataset()
    finding_site.RelationshipType = "CONTAINS"
    finding_site.ValueType = "CODE"
    finding_site.ConceptNameCodeSequence = [Dataset()]
    fscn = finding_site.ConceptNameCodeSequence[0]
    fscn.CodeValue, fscn.CodingSchemeDesignator, fscn.CodeMeaning = FINDING_SITE_CODE
    # Mark the finding site as \"hidden\" / unannotated, matching the template SR.
    hidden_code = _make_code(*HIDDEN_UNANNOTATED_CODE)
    finding_site.ConceptCodeSequence = [hidden_code]

    num = Dataset()
    num.RelationshipType = "CONTAINS"
    num.ValueType = "NUM"
    num.ConceptNameCodeSequence = [Dataset()]
    ncn = num.ConceptNameCodeSequence[0]
    ncn.CodeValue, ncn.CodingSchemeDesignator, ncn.CodeMeaning = POINT_CODE

    scoord = Dataset()
    scoord.RelationshipType = "INFERRED FROM"
    scoord.ValueType = "SCOORD"
    scoord.GraphicType = "POINT"
    scoord.GraphicData = [float(coords[0]), float(coords[1])]

    image_item = Dataset()
    image_item.RelationshipType = "SELECTED FROM"
    image_item.ValueType = "IMAGE"
    image_item.ReferencedSOPSequence = [copy.deepcopy(image_reference)]
    scoord.ContentSequence = [image_item]
    num.ContentSequence = [scoord]

    finding = Dataset()
    finding.RelationshipType = "CONTAINS"
    finding.ValueType = "CODE"
    finding.ConceptNameCodeSequence = [Dataset()]
    fcn = finding.ConceptNameCodeSequence[0]
    fcn.CodeValue, fcn.CodingSchemeDesignator, fcn.CodeMeaning = FINDING_CODE
    freetext = Dataset()
    freetext.CodeValue, freetext.CodingSchemeDesignator = CORNERSTONE_FREETEXT
    freetext.CodeMeaning = label
    finding.ConceptCodeSequence = [freetext]

    # Order of items in ContentSequence matches the template SR:
    # [0] Tracking Identifier (TEXT)
    # [1] Tracking UID (UIDREF)
    # [2] Finding Site (CODE) with HIDDEN/99MEDICALVIEWER
    # [3] Center (NUM) with SCOORD point
    # [4] Finding label (CODE) with CORNERSTONEFREETEXT
    group.ContentSequence.extend([finding_site, num, finding])
    return group


def _make_imaging_measurements(
    labels: Sequence[Tuple[str, Tuple[float, float]]],
    image_reference: Dataset,
) -> Dataset:
    container = Dataset()
    container.RelationshipType = "CONTAINS"
    container.ValueType = "CONTAINER"
    container.ConceptNameCodeSequence = [Dataset()]
    cn = container.ConceptNameCodeSequence[0]
    cn.CodeValue, cn.CodingSchemeDesignator, cn.CodeMeaning = IMAGING_MEASUREMENTS_CODE
    container.ContinuityOfContent = "SEPARATE"
    container.ContentSequence = [
        _make_measurement_group(label, coords, image_reference)
        for label, coords in labels
    ]
    return container


def _default_placeholder_coords(idx: int) -> Tuple[float, float]:
    col = idx % 5
    row = idx // 5
    return (50.0 + col * 25.0, 50.0 + row * 25.0)


def _parse_point_labels(path: Path) -> List[Tuple[str, Tuple[float, float]]]:
    labels: List[Tuple[str, Tuple[float, float]]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            raw = line.strip()
            if not raw or raw.startswith("#"):
                continue
            parts = raw.replace(",", " ").split()
            label = parts[0]
            if not label.startswith("pt-"):
                continue
            coords: List[float] = []
            for token in parts[1:]:
                try:
                    coords.append(float(token))
                except ValueError:
                    continue
            if len(coords) >= 2:
                xy = (coords[0], coords[1])
            else:
                xy = _default_placeholder_coords(len(labels))
            labels.append((label, xy))
    if not labels:
        raise ValueError(f"No point labels found in {path}")
    return labels


def _normalize_label(raw: str) -> str:
    s = raw.strip()
    if not s:
        return s
    return s if s.startswith("pt-") else f"pt-{s}"


def _view_from_subfolder(subfolder: str) -> str:
    up = subfolder.upper()
    if "LAT" in up:
        return "LAT"
    if "AP" in up:
        return "AP"
    return "UNKNOWN"


def _expand_spec_token(token: str, view: str) -> List[str]:
    """
    Expand a single spec token from the XLSX into one or more point labels.

    Rules:
    - `*-c` => one point
    - `*-edges` => two points (LAT: ant-sup, pos-sup; AP: left-sup, right-sup)
    - `*-corners` => four points (LAT: ant/pos x sup/inf; AP: left/right x sup/inf)
    - `FH1-circle` / `FH2-circle` => mapped to `pt-FH-1` / `pt-FH-2` for consistency
    - otherwise => one point with `pt-` prefix
    """
    t = token.strip()
    if not t:
        return []

    # Special-case femoral head circles from the template
    if t.upper() in {"FH1-CIRCLE", "FH2-CIRCLE"}:
        n = "1" if t.upper().startswith("FH1") else "2"
        return [f"pt-FH-{n}"]

    if t.lower().endswith("-edges"):
        base = t[: -len("-edges")]
        if view == "AP":
            suffixes = ["left-sup", "right-sup"]
        else:  # LAT and unknown fall back to LAT convention
            suffixes = ["ant-sup", "pos-sup"]
        return [f"pt-{base}-{s}" for s in suffixes]

    if t.lower().endswith("-corners"):
        base = t[: -len("-corners")]
        if view == "AP":
            suffixes = ["left-sup", "left-inf", "right-sup", "right-inf"]
        else:  # LAT and unknown fall back to LAT convention
            suffixes = ["ant-sup", "ant-inf", "pos-sup", "pos-inf"]
        return [f"pt-{base}-{s}" for s in suffixes]

    return [_normalize_label(t)]


def _iter_unique(items: Iterable[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for item in items:
        if not item:
            continue
        if item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def _parse_xlsx_specs(
    xlsx_path: Path,
    sheet_name: str = DEFAULT_XLSX_SHEET,
) -> Dict[Tuple[str, str], List[str]]:
    """
    Parse the refdata-template XLSX sheet into a mapping:
        (case_folder, image_subfolder) -> list of pt-* labels
    """
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for --input-xlsx. "
            "Install it into your venv: `python -m pip install openpyxl`"
        ) from exc

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {xlsx_path}. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    header_row = 1
    headers = [
        str(ws.cell(header_row, c).value).strip()
        if ws.cell(header_row, c).value is not None
        else ""
        for c in range(1, ws.max_column + 1)
    ]
    if len(headers) < 2:
        raise ValueError(f"Unexpected header in {xlsx_path}:{sheet_name}")

    folder_col = 1
    subfolder_col = 2

    result: Dict[Tuple[str, str], List[str]] = {}
    last_folder: Optional[str] = None

    for r in range(2, ws.max_row + 1):
        folder_val = ws.cell(r, folder_col).value
        subfolder_val = ws.cell(r, subfolder_col).value

        folder = (str(folder_val).strip() if folder_val is not None else "") or ""
        if folder:
            last_folder = folder
        elif last_folder:
            folder = last_folder

        subfolder = (
            str(subfolder_val).strip() if subfolder_val is not None else ""
        ).strip()

        if not folder or not subfolder:
            continue

        view = _view_from_subfolder(subfolder)

        tokens: List[str] = []
        for c in range(3, ws.max_column + 1):
            cell = ws.cell(r, c).value
            if cell is None:
                continue
            raw = str(cell).strip()
            if not raw:
                continue
            # Comma-separated tokens like: "S1-c, S1-edges"
            for part in raw.replace("\n", " ").split(","):
                part = part.strip()
                if not part:
                    continue
                tokens.extend(_expand_spec_token(part, view))

        key = (folder, subfolder)
        if key not in result:
            result[key] = []
        result[key].extend(tokens)

    # De-duplicate while preserving order
    for key, labels in list(result.items()):
        result[key] = _iter_unique(labels)
        if not result[key]:
            del result[key]
    return result


def _build_image_reference(ct_ds: Dataset) -> Dataset:
    ref = Dataset()
    ref.ReferencedSOPClassUID = ct_ds.SOPClassUID
    ref.ReferencedSOPInstanceUID = ct_ds.SOPInstanceUID
    return ref


def _update_patient_and_study(sr_ds: Dataset, ct_ds: Dataset) -> None:
    for attr in [
        "PatientName",
        "PatientID",
        "PatientBirthDate",
        "PatientSex",
        "StudyInstanceUID",
        "StudyID",
        "StudyDate",
        "StudyTime",
        "AccessionNumber",
        "ReferringPhysicianName",
    ]:
        if hasattr(ct_ds, attr):
            setattr(sr_ds, attr, getattr(ct_ds, attr))


def _update_uids(sr_ds: Dataset) -> None:
    new_series_uid = generate_uid()
    new_instance_uid = generate_uid()
    sr_ds.SeriesInstanceUID = new_series_uid
    sr_ds.SOPInstanceUID = new_instance_uid
    if hasattr(sr_ds, "file_meta"):
        sr_ds.file_meta.MediaStorageSOPInstanceUID = new_instance_uid


def _update_dates(sr_ds: Dataset) -> None:
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H%M%S")
    sr_ds.ContentDate = date_str
    sr_ds.ContentTime = time_str
    sr_ds.SeriesDate = date_str
    sr_ds.SeriesTime = time_str
    if not getattr(sr_ds, "SeriesDescription", "").startswith("Generated"):
        sr_ds.SeriesDescription = "Generated SR from CT-axi-postop.txt"


def _build_evidence_sequence(ct_ds: Dataset) -> List[Dataset]:
    sop = Dataset()
    sop.ReferencedSOPClassUID = ct_ds.SOPClassUID
    sop.ReferencedSOPInstanceUID = ct_ds.SOPInstanceUID

    series = Dataset()
    series.SeriesInstanceUID = ct_ds.SeriesInstanceUID
    series.ReferencedSOPSequence = [sop]

    study = Dataset()
    study.StudyInstanceUID = ct_ds.StudyInstanceUID
    study.ReferencedSeriesSequence = [series]
    return [study]


def _build_content_sequence(
    labels: Sequence[Tuple[str, Tuple[float, float]]],
    image_reference: Dataset,
    observer_name: str,
) -> List[Dataset]:
    content_items = [
        _make_language_item(),
        _make_observer_item(observer_name),
        _make_procedure_item(),
        _make_image_library(image_reference),
        _make_imaging_measurements(labels, image_reference),
    ]
    return content_items


def _find_first_ct_instance(ct_folder: Path) -> Path:
    # Prefer direct files, but also support one level of nesting.
    for path in sorted(ct_folder.iterdir()):
        if path.is_file():
            return path
    for path in sorted(ct_folder.iterdir()):
        if path.is_dir():
            for sub in sorted(path.iterdir()):
                if sub.is_file():
                    return sub
    raise FileNotFoundError(f"No DICOM instances found in {ct_folder}")


def _labels_with_placeholder_coords(labels: Sequence[str]) -> List[Tuple[str, Tuple[float, float]]]:
    return [(label, _default_placeholder_coords(i)) for i, label in enumerate(labels)]


def main() -> None:
    scripts_dir = Path(__file__).resolve().parent
    default_txt = scripts_dir / "CT-axi-postop.txt"
    default_template = scripts_dir / "SR Saved report" / "SR000001.dcm"
    default_ct = scripts_dir / "CT-axi-postop"
    default_output = scripts_dir / "SR Saved report" / "SR_from_txt.dcm"
    default_xlsx = scripts_dir / "data_J2-4452.xlsx"
    default_dicom_root = scripts_dir.parent / "testdata"

    parser = argparse.ArgumentParser(
        description="Create an SR report from CT-axi-postop annotations."
    )
    in_group = parser.add_mutually_exclusive_group()
    in_group.add_argument(
        "--input-txt",
        type=Path,
        default=default_txt,
        help="Plain text label list (legacy mode).",
    )
    in_group.add_argument(
        "--input-xlsx",
        type=Path,
        help="XLSX refdata template describing which annotations should exist.",
    )
    parser.add_argument(
        "--xlsx-sheet",
        default=DEFAULT_XLSX_SHEET,
        help=f"Sheet name for --input-xlsx (default: {DEFAULT_XLSX_SHEET}).",
    )
    parser.add_argument("--template", type=Path, default=default_template)
    parser.add_argument(
        "--ct-folder",
        type=Path,
        default=default_ct,
        help="DICOM folder for --input-txt mode.",
    )
    parser.add_argument(
        "--dicom-root",
        type=Path,
        default=default_dicom_root,
        help="Root directory that contains case folders (used for --input-xlsx mode).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=default_output,
        help="Output SR file for --input-txt mode, or an output directory for --input-xlsx mode.",
    )
    parser.add_argument(
        "--observer-name",
        default="unknown^unknown",
        help="Person observer name encoded as CARET-delimited PN.",
    )
    parser.add_argument(
        "--case",
        help="Optional case folder filter for --input-xlsx mode (e.g. obv_case001).",
    )
    parser.add_argument(
        "--subfolder",
        help="Optional subfolder filter for --input-xlsx mode (e.g. CT-axi-postop).",
    )
    args = parser.parse_args()

    template_ds = pydicom.dcmread(str(args.template))

    if args.input_xlsx:
        # XLSX mode: generate one SR per case/subfolder combination.
        xlsx_path: Path = args.input_xlsx
        specs = _parse_xlsx_specs(xlsx_path, sheet_name=args.xlsx_sheet)

        if args.case:
            specs = {k: v for k, v in specs.items() if k[0] == args.case}
        if args.subfolder:
            specs = {k: v for k, v in specs.items() if k[1] == args.subfolder}

        out_base: Path = args.output
        out_base.mkdir(parents=True, exist_ok=True)

        created = 0
        for (case_folder, image_subfolder), label_list in sorted(specs.items()):
            dicom_folder = args.dicom_root / case_folder / image_subfolder
            if not dicom_folder.exists():
                print(f"WARNING: Missing DICOM folder, skipping: {dicom_folder}")
                continue

            first_path = _find_first_ct_instance(dicom_folder)
            ref_ds = pydicom.dcmread(str(first_path))
            labels = _labels_with_placeholder_coords(label_list)

            sr_ds = copy.deepcopy(template_ds)
            _update_patient_and_study(sr_ds, ref_ds)
            _update_uids(sr_ds)
            _update_dates(sr_ds)
            sr_ds.SeriesNumber = str(getattr(ref_ds, "SeriesNumber", 1))
            sr_ds.InstanceNumber = "1"
            sr_ds.Modality = "SR"

            sr_ds.SeriesDescription = f"Generated SR from XLSX: {case_folder}/{image_subfolder}"
            sr_ds.CurrentRequestedProcedureEvidenceSequence = _build_evidence_sequence(ref_ds)

            image_ref = _build_image_reference(ref_ds)
            sr_ds.ContentSequence = _build_content_sequence(
                labels, image_ref, observer_name=args.observer_name
            )

            out_path = out_base / f"{case_folder}__{image_subfolder}.dcm"
            sr_ds.save_as(str(out_path), write_like_original=False)
            created += 1
            print(
                f"Created SR with {len(labels)} measurement groups -> {out_path} "
                f"(referencing {first_path.name})"
            )

        if created == 0:
            raise RuntimeError("No SR files created (filters too strict or missing DICOM folders).")
        return

    # TXT mode (legacy, single SR)
    input_txt: Path = args.input_txt
    if not input_txt.exists() and default_xlsx.exists():
        # Backwards-friendly: if the legacy default file isn't present, keep error message useful.
        print(f"NOTE: {input_txt} not found. If you meant XLSX mode, use --input-xlsx {default_xlsx}")

    labels = _parse_point_labels(input_txt)
    ct_path = _find_first_ct_instance(args.ct_folder)
    ct_ds = pydicom.dcmread(str(ct_path))

    sr_ds = copy.deepcopy(template_ds)
    _update_patient_and_study(sr_ds, ct_ds)
    _update_uids(sr_ds)
    _update_dates(sr_ds)
    sr_ds.SeriesNumber = str(getattr(ct_ds, "SeriesNumber", 1))
    sr_ds.InstanceNumber = "1"
    sr_ds.Modality = "SR"

    sr_ds.SeriesDescription = f"Generated SR from {input_txt.name}"
    sr_ds.CurrentRequestedProcedureEvidenceSequence = _build_evidence_sequence(ct_ds)

    image_ref = _build_image_reference(ct_ds)
    sr_ds.ContentSequence = _build_content_sequence(
        labels, image_ref, observer_name=args.observer_name
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    sr_ds.save_as(str(args.output), write_like_original=False)
    print(
        f"Created SR with {len(labels)} measurement groups -> {args.output} "
        f"(referencing {ct_path.name})"
    )


if __name__ == "__main__":
    main()
